#!/usr/bin/env python3
"""
xlsx_to_json.py — 엑셀 원본 → JSON 변환

지원하는 산출물:
  1) applicant_compare       (모집단위별 2025 vs 2026 지원자 수, 전형별 분포)
  2) registration_2026       (2026 모집단위별 합격구분 집계: 최초합격/충원합격/등록률)
  3) grade_heatmap_2025      (작성 시트 → 석차등급 분포도)
  4) grade_heatmap_2026      (샘플 시트 → 석차등급 분포도)

보안: 지원자 시트는 학생 raw 데이터(수험번호·출신고교·소재지)이므로
      절대 그대로 출력하지 않고, group-by 집계값만 산출한다.
"""
from __future__ import annotations
import argparse, json, sys, re
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl


# -------------------- 유틸 --------------------

def norm(s):
    """문자열 정규화: 줄바꿈·연속 공백을 단일 공백으로, 양끝 trim."""
    if s is None: return ''
    return re.sub(r'\s+', ' ', str(s)).strip()

def is_num(v):
    return isinstance(v, (int, float)) and not isinstance(v, bool)


# -------------------- 1) 지원자 시트 집계 --------------------

def aggregate_applicants(path: Path, sheet_name: str) -> dict:
    """
    `지원자` 시트의 (단과대학, 모집단위, 전형구분) 열을 group-by count.
    개인정보(수험번호·출신고교)는 절대 출력하지 않는다.
    """
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if sheet_name not in wb.sheetnames:
        raise SystemExit(f'sheet not found: "{sheet_name}" in {path}')
    ws = wb[sheet_name]

    # 헤더 1행에서 필요한 컬럼 인덱스 확보
    header = {}
    for c, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)), start=1):
        if cell is not None:
            header[norm(cell)] = c
    needed = ['단과대학', '모집단위', '전형구분', '모집구분']
    missing = [h for h in needed if h not in header]
    if missing:
        raise SystemExit(f'헤더 누락: {missing} in "{sheet_name}". 실제 헤더: {list(header)[:30]}')

    col_college = header['단과대학']
    col_unit    = header['모집단위']
    col_form    = header['전형구분']
    col_kind    = header['모집구분']
    col_admit   = header.get('합격구분')   # 2026 지원자 전용 (없을 수도 있음)

    by_unit_total      = Counter()                # (단과대학, 모집단위) → 지원자 수
    by_unit_form       = defaultdict(Counter)     # (단과대학, 모집단위) → {전형구분: 수}
    by_unit_admit      = defaultdict(Counter)     # 모집단위 → {합격구분: 수}
    by_form_total      = Counter()                # 전형구분 → 총 지원자 수

    for row in ws.iter_rows(min_row=2, values_only=True):
        kind = norm(row[col_kind - 1])
        if not kind or '수시' not in kind:
            continue
        college = norm(row[col_college - 1])
        unit    = norm(row[col_unit    - 1])
        form    = norm(row[col_form    - 1])
        if not unit:
            continue
        key = (college, unit)
        by_unit_total[key] += 1
        by_unit_form[key][form] += 1
        by_form_total[form] += 1
        if col_admit:
            admit = norm(row[col_admit - 1])
            by_unit_admit[unit][admit] += 1

    return {
        'sheet': sheet_name,
        'unit_total':  {f'{c}|{u}': n for (c, u), n in by_unit_total.items()},
        'unit_form':   {f'{c}|{u}': dict(d) for (c, u), d in by_unit_form.items()},
        'unit_admit':  {u: dict(d) for u, d in by_unit_admit.items()},
        'form_total':  dict(by_form_total),
    }


# -------------------- 2) 석차등급 분포 시트 --------------------

def extract_grade_heatmap(path: Path, sheet_name: str, year: int, table_id: str) -> dict:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
    if sheet_name not in wb.sheetnames:
        raise SystemExit(f'sheet not found: "{sheet_name}" in {path}')
    ws = wb[sheet_name]

    # 1행: 제목, 2행: 헤더, 3행~: 데이터
    title_cell = ws.cell(1, 1).value or ''
    title = norm(title_cell).lstrip('◦').strip() or f'{year} 석차등급 분포'

    headers = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(2, c).value
        if v is None:
            break
        headers.append(norm(v))

    rows = []
    for r in range(3, ws.max_row + 1):
        first = ws.cell(r, 1).value
        if first is None or norm(first) == '':
            continue
        first_norm = norm(first)
        # '계' 행과 메모/주석 행은 제외
        if first_norm in ('계', '소계', '합계') or first_norm.startswith('※') or first_norm.startswith('*'):
            continue
        row = []
        for c in range(1, len(headers) + 1):
            v = ws.cell(r, c).value
            if v == '-' or v is None:
                row.append(0)
            elif is_num(v):
                row.append(round(float(v), 2))
            else:
                row.append(norm(v))
        rows.append(row)

    # 행 단위 백분율로 변환 (마지막 컬럼 '계' 기준) — 히트맵 색상 사양 호환
    pct_headers = headers[:-1] if headers and headers[-1] == '계' else headers
    pct_rows = []
    for row in rows:
        unit = row[0]
        try:
            counts = [v if isinstance(v, (int, float)) else 0 for v in row[1:]]
        except Exception:
            continue
        # 마지막이 '계'면 분리
        if headers and headers[-1] == '계':
            total = counts[-1] if counts else 0
            grade_counts = counts[:-1]
        else:
            total = sum(counts)
            grade_counts = counts
        if not total:
            pct_rows.append([unit] + [0.0 for _ in grade_counts])
            continue
        pct_rows.append([unit] + [round(g / total * 100, 1) for g in grade_counts])

    return {
        'id': table_id,
        'year': year,
        'category': 'I. 수시 모집 지원 현황',
        'subcategory': '3. 학생부 석차등급 분포',
        'title': f'{year} 등록자 석차등급 분포(지역교과)',
        'subtitle': f'단위: % (각 행 합계 100%) — 색상: <15% 노랑 / 15~40% 옅은 주황 / >40% 짙은 주황 · 원본: "{title}"',
        'headers': ['모집단위'] + list(pct_headers[1:]),
        'rows': pct_rows,
        'totals': None,
        'chart_hint': 'heatmap_grade',
        'meta': {
            'source_file': path.name,
            'source_sheet': sheet_name,
            'updated_at': '2026-05-06',
            'status': 'final',
            'note': '백분율은 자동 산출(원본은 절대 인원). 1차 등록자 기준'
        },
        # 검증용 raw counts도 포함 (감사용, 페이지에서는 미사용)
        '_raw_counts': {'headers': headers, 'rows': rows}
    }


# -------------------- 3) 비교 산출 (2025 vs 2026) --------------------

def build_compare(agg25: dict, agg26: dict) -> dict:
    """모집단위별 2025·2026 지원자 수 비교 + 변동률."""
    keys = sorted(set(agg25['unit_total']) | set(agg26['unit_total']))
    rows = []
    for key in keys:
        college, unit = key.split('|', 1)
        v25 = agg25['unit_total'].get(key, 0)
        v26 = agg26['unit_total'].get(key, 0)
        delta = v26 - v25
        pct = round((delta / v25 * 100), 1) if v25 else None
        rows.append([college, unit, v25, v26, delta, pct if pct is not None else '-'])
    rows.sort(key=lambda r: (r[0], r[1]))

    return {
        'id': 'COMPARE-APPLICANTS',
        'category': '비교',
        'subcategory': '모집단위별 지원자 수 비교',
        'title': '2025 vs 2026 모집단위별 지원자 수',
        'subtitle': '단위: 명 · 변동률 = (2026 − 2025) / 2025 × 100%',
        'headers': ['단과대학', '모집단위', '2025', '2026', '변동(±)', '변동률(%)'],
        'rows': rows,
        'totals': [
            '합계', '',
            sum(r[2] for r in rows),
            sum(r[3] for r in rows),
            sum(r[3] for r in rows) - sum(r[2] for r in rows),
            round((sum(r[3] for r in rows) - sum(r[2] for r in rows))
                  / max(1, sum(r[2] for r in rows)) * 100, 1)
        ],
        'chart_hint': 'bar_yoy_applicants',
        'meta': {
            'source_file': '최근 2년 모집인원 및 지원자.xlsx',
            'source_sheets': [agg25['sheet'], agg26['sheet']],
            'updated_at': '2026-05-06',
            'status': 'final',
            'aggregation': 'group_by(단과대학, 모집단위) on 모집구분 contains "수시"'
        }
    }


# -------------------- 4) 2026 합격구분 → 등록 통계 --------------------

# '합격구분' 값을 표준 카테고리에 매핑
ADMIT_NORMALIZE = {
    '최초합격':      '최초합격',
    '외국인합격':    '최초합격',     # 외국인 전형 최초합격
    '추가합격':      '충원합격',
    '충원합격':      '충원합격',
    '1차충원': '충원합격', '2차충원': '충원합격', '3차충원': '충원합격',
    '4차충원': '충원합격', '5차충원': '충원합격', '6차충원': '충원합격',
    '7차충원': '충원합격', '8차충원': '충원합격',
    '예비순위':      '예비순위',
    '예비등록':      '예비순위',
    '1단계 불합격':  '1단계 불합격',
    '2단계 불합격':  '2단계 불합격',
    '불합격':        '불합격',
    '전형제외':      '전형제외',
    '결시':          '전형제외',
    '등록':          '등록',
    '미등록':        '미등록',
    '':              '미상',
}

def build_registration_2026(agg26: dict) -> dict:
    """단과대학·모집단위 단위 합격구분 집계."""
    rows = []
    unit_to_college = {}
    for key in agg26['unit_total']:
        college, unit = key.split('|', 1)
        unit_to_college.setdefault(unit, college)

    for unit, admit_counts in sorted(agg26['unit_admit'].items()):
        cat = Counter()
        for raw, n in admit_counts.items():
            cat[ADMIT_NORMALIZE.get(raw, '기타')] += n
        college = unit_to_college.get(unit, '')
        first    = cat['최초합격']
        topup    = cat['충원합격']
        reserve  = cat['예비순위']
        stage1f  = cat['1단계 불합격']
        excluded = cat['전형제외'] + cat['불합격'] + cat['2단계 불합격']
        passed = first + topup
        applicants = sum(admit_counts.values())
        pass_rate = round(passed / applicants * 100, 1) if applicants else 0.0
        rows.append([college, unit, applicants, first, topup, reserve, stage1f, excluded, passed, pass_rate])
    rows.sort(key=lambda r: (r[0], r[1]))

    return {
        'id': 'IV-1-01',
        'year': 2026,
        'category': 'IV. 수시모집 합격 및 등록 현황',
        'subcategory': '1. 모집단위별 합격구분 집계',
        'title': '2026 모집단위별 합격구분 집계',
        'subtitle': '지원자 시트의 합격구분을 group-by 집계 · 충원합격은 1~N차 충원·추가합격 합산 · 외국인합격은 최초합격에 포함',
        'headers': ['단과대학', '모집단위', '지원자', '최초합격', '충원합격',
                    '예비순위', '1단계 불합격', '전형제외·불합격', '합격(최초+충원)', '합격률(%)'],
        'rows': rows,
        'totals': [
            '합계', '',
            sum(r[2] for r in rows), sum(r[3] for r in rows),
            sum(r[4] for r in rows), sum(r[5] for r in rows),
            sum(r[6] for r in rows), sum(r[7] for r in rows),
            sum(r[8] for r in rows),
            round(sum(r[8] for r in rows) / max(1, sum(r[2] for r in rows)) * 100, 1)
        ],
        'chart_hint': 'bar_pass_rate',
        'meta': {
            'source_file': '최근 2년 모집인원 및 지원자.xlsx',
            'source_sheet': '2026지원자',
            'updated_at': '2026-05-06',
            'status': 'final',
            'note': '등록률(=등록자/모집인원)이 아닌 합격률(=합격자/지원자)임. 정원 매칭은 hwpx 변환 후 통합 예정'
        },
        '_raw_admit_categories': {u: dict(d) for u, d in agg26['unit_admit'].items()}
    }


# -------------------- CLI --------------------

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--input',  default='최근 2년 모집인원 및 지원자.xlsx')
    ap.add_argument('--out-dir', default='data')
    args = ap.parse_args()

    src = Path(args.input)
    if not src.exists():
        sys.exit(f'not found: {src}')

    out_dir = Path(args.out_dir)
    (out_dir / '2025').mkdir(parents=True, exist_ok=True)
    (out_dir / '2026').mkdir(parents=True, exist_ok=True)
    (out_dir / 'compare').mkdir(parents=True, exist_ok=True)

    # 1) 지원자 시트 집계 (2025, 2026)
    print('[1/4] 지원자 시트 집계 (개인정보 제외, group-by 카운트)...')
    agg25 = aggregate_applicants(src, '2025 지원자')
    agg26 = aggregate_applicants(src, '2026지원자')   # 시트명에 공백 없음
    print(f'      2025 모집단위 수: {len(agg25["unit_total"])}, 전형 수: {len(agg25["form_total"])}')
    print(f'      2026 모집단위 수: {len(agg26["unit_total"])}, 전형 수: {len(agg26["form_total"])}')

    # 2) 비교 산출
    print('[2/4] 2025 vs 2026 비교 산출...')
    compare = build_compare(agg25, agg26)
    out_path = out_dir / 'compare' / '2025_vs_2026_applicants.json'
    out_path.write_text(json.dumps(compare, ensure_ascii=False, indent=2), encoding='utf-8')
    print(f'      → {out_path}')

    # 3) 2026 합격구분 → 등록 산출
    if agg26['unit_admit']:
        print('[3/4] 2026 합격구분 집계 (IV-1-01)...')
        reg = build_registration_2026(agg26)
        out_path = out_dir / '2026' / 'IV-1-01_최종등록.json'
        out_path.write_text(json.dumps(reg, ensure_ascii=False, indent=2), encoding='utf-8')
        print(f'      → {out_path}  ({len(reg["rows"])}개 모집단위)')
    else:
        print('[3/4] 합격구분 컬럼이 없어 건너뜀')

    # 4) 석차등급 히트맵 (작성=2025등록자, 샘플=2026등록자)
    print('[4/4] 석차등급 분포(히트맵) 추출...')
    h25 = extract_grade_heatmap(src, '작성', 2025, 'I-3-01')
    h26 = extract_grade_heatmap(src, '샘플', 2026, 'I-3-01')
    (out_dir / '2025' / 'I-3-01_석차등급분포.json').write_text(
        json.dumps(h25, ensure_ascii=False, indent=2), encoding='utf-8')
    (out_dir / '2026' / 'I-3-01_석차등급분포.json').write_text(
        json.dumps(h26, ensure_ascii=False, indent=2), encoding='utf-8')
    print(f'      → 2025 히트맵: {len(h25["rows"])}행')
    print(f'      → 2026 히트맵: {len(h26["rows"])}행')

    print('\n완료.')


if __name__ == '__main__':
    main()
